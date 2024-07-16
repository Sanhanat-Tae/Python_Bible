import openpyxl as xl

file = xl.load_workbook("custexcel.xlsx")
sheet = file.active

cell1 = sheet["A2"]
print("Cell A2 = "+cell1.value)
cell2 = sheet.cell(row=6,column=1)
print("\nCell A6 = "+cell2.value)

cells = sheet['A1': 'B6']
print("\n[Customer name list]")
for c1, c2 in cells:
    print(c1.value+ " " + c2.value)


print("\n[Last customer]")
for row in sheet.iter_rows(min_row=6, max_col=5, max_row=6):
    for cell3 in row:
        print(cell3.value)
        
print("\n[Customer email]")
for row in sheet.iter_cols(min_row=2, min_col=4, max_row=6, max_col=4):
    for cell4 in row:
        print(cell4.value)